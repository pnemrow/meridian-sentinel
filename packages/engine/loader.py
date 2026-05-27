"""
Spreadsheet loader — reads entities from .xlsx.

Ported from sayari_ground_truth.py. Column detection is hint-based so
the loader survives minor header renaming.
"""
from __future__ import annotations

from pathlib import Path

from .types import InputEntity

COLUMN_HINTS = {
    "name":       ["name", "entity", "company", "supplier", "vendor", "counterparty"],
    "address":    ["address", "street", "location"],
    "country":    ["country", "nation", "jurisdiction"],
    "type":       ["type", "entity_type", "kind"],
    "identifier": ["identifier", "id_number", "reg", "tax", "duns"],
}


def _detect_columns(header: list) -> dict[str, int]:
    mapping: dict[str, int] = {}
    lowered = [(i, str(h).strip().lower()) for i, h in enumerate(header) if h is not None]
    for field_name, hints in COLUMN_HINTS.items():
        for i, h in lowered:
            if any(hint in h for hint in hints):
                mapping[field_name] = i
                break
    return mapping


def load_entities_from_xlsx(path: Path, sheet: str = "list_1") -> list[InputEntity]:
    """Read entity list from .xlsx. Auto-detects columns; 'name' is required."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.active
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    cols = _detect_columns(header)
    if "name" not in cols:
        raise ValueError(
            f"Could not find a name column in {header!r}. "
            "Rename the column to include 'name' or edit COLUMN_HINTS."
        )

    def cell(values: tuple, key: str) -> str | None:
        idx = cols.get(key)
        if idx is None or idx >= len(values):
            return None
        v = values[idx]
        return str(v).strip() if v not in (None, "") else None

    entities: list[InputEntity] = []
    for n, values in enumerate(rows, start=2):
        name = cell(values, "name")
        if not name:
            continue
        entities.append(
            InputEntity(
                row=n,
                name=name,
                address=cell(values, "address"),
                country=cell(values, "country"),
                type=cell(values, "type"),
                identifier=cell(values, "identifier"),
            )
        )
    return entities
