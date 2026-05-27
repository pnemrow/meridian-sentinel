#!/usr/bin/env python3
"""
sayari_ground_truth.py
======================

A *deterministic* reference pipeline for the Sayari FDE technical exercise.

Purpose
-------
This script is intentionally NOT an AI agent. It is a small, readable,
reproducible pipeline whose job is to be a "source of truth" you can trust
and that you can diff your Replit agent's output against. Same input ->
same output, every time, with every value traceable back to a specific
Sayari API response field.

What it does (Scenario 2: Analytics Report, with the profile-retrieval core
that Scenario 1 would also need):

    1. Read a list of entities from an .xlsx (auto-detects name/address/
       country/type columns).
    2. RESOLVE each row to its best-match Sayari entity_id (resolution
       endpoint), logging the match score so resolution quality is auditable
       -- this is the #1 place these pipelines silently go wrong.
    3. FETCH the full profile for each matched entity (get_entity endpoint),
       saving the *raw* JSON to disk as an audit trail.
    4. EXTRACT a flat, traceable record per entity (id, type, countries,
       sanctioned, pep, risk factors, relationship counts, source count, and
       the web-UI url so a human can eyeball it).
    5. AGGREGATE macro-level insights (country breakdown, % sanctioned,
       % PEP, risk-factor frequency, entity-type mix).
    6. WRITE: raw/<id>.json, resolution_log.csv, entities.csv, summary.json

Design principles (these map directly to the rubric: "correct use of the
Sayari API", "readability, simplicity"):
    * No LLM in the loop  -> output is reproducible and verifiable.
    * Raw responses saved  -> every number is traceable to a source field.
    * Resolution is logged -> you can see *which* entity each row matched and
      how confident the match was, instead of trusting it blindly.
    * Field extraction is DEFENSIVE (tries multiple key names) so the script
      survives minor schema differences; confirm exact names against the
      raw/<id>.json files once you've run it live.

Usage
-----
    pip install sayari openpyxl
    export SAYARI_CLIENT_ID=...           # from the setup email
    export SAYARI_CLIENT_SECRET=...
    python sayari_ground_truth.py --input entities.xlsx --out ./output

    # peek at one entity end-to-end without writing files:
    python sayari_ground_truth.py --input entities.xlsx --limit 1 --verbose
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import time
from collections import Counter
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Any, Iterable

from unidecode import unidecode

# ---------------------------------------------------------------------------
# Small generic helpers
# ---------------------------------------------------------------------------

def to_dict(obj: Any) -> Any:
    """Convert an SDK response object into a plain dict/list.

    The Sayari SDK returns Pydantic models. We normalise everything to plain
    Python so (a) extraction is uniform and (b) we can dump the raw payload to
    JSON as an audit trail. Works across Pydantic v1/v2 and dataclasses.
    """
    if obj is None or isinstance(obj, (str, int, float, bool)):
        return obj
    if isinstance(obj, dict):
        return {k: to_dict(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [to_dict(v) for v in obj]
    for attr in ("model_dump", "dict"):  # Pydantic v2, then v1
        fn = getattr(obj, attr, None)
        if callable(fn):
            try:
                return to_dict(fn())
            except Exception:
                pass
    if hasattr(obj, "__dict__"):
        return {k: to_dict(v) for k, v in vars(obj).items() if not k.startswith("_")}
    return str(obj)


def deep_get(d: Any, *keys: str, default: Any = None) -> Any:
    """Safely walk nested dict keys, returning ``default`` if any are missing."""
    cur = d
    for k in keys:
        if isinstance(cur, dict) and k in cur:
            cur = cur[k]
        else:
            return default
    return cur


def first_present(d: dict, *keys: str, default: Any = None) -> Any:
    """Return the value of the first key that exists in ``d``.

    Used because the API/SDK may name a field slightly differently than the
    docs snippet (e.g. ``entity_id`` vs ``id``). Confirm against raw JSON.
    """
    for k in keys:
        if isinstance(d, dict) and d.get(k) is not None:
            return d[k]
    return default


# ---------------------------------------------------------------------------
# Step 1: read entities from the spreadsheet
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Retry alternates for entities that fail primary resolution
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Verified parent entity IDs — applied before any API call
# ---------------------------------------------------------------------------
# These were confirmed by fetching all resolution candidates, comparing degree
# (relationship count) and translated_label, and choosing the entity with the
# highest degree that is the publicly-known parent/principal entity.
#
# To re-verify: delete an entry, re-run, inspect output/raw/row{n}_resolution.json,
# check degree via get_entity for each candidate, and restore the correct ID.
#
# Entity                 | Pinned ID                  | Degree  | Why pinned
# Sberbank               | OWwtbp9y51OcLHJQakLaMw     | 59,632  | [0] was LLC Sberbank Service (degree 780)
# VTB Bank               | dy-rh2g0QtzUN_jC_e9S_A     | 15,509  | [0] was VTB Capital Holdings (degree 208)
# Transneft              | 9-IuyJoA08bELHrSY3mXXA     |  1,526  | [0] was TRANSNEFT (degree 464, stub)
PINNED_IDS: dict[str, str] = {
    "Sberbank":  "OWwtbp9y51OcLHJQakLaMw",   # PJSC Sberbank of Russia
    "VTB Bank":  "dy-rh2g0QtzUN_jC_e9S_A",   # VTB Bank PAO
    "Transneft": "9-IuyJoA08bELHrSY3mXXA",   # Transneft PAO
}

# Keys are the exact input names that return no candidates on first try.
# Values are ordered list of alternate names to try, stopping at first hit.
RETRY_NAMES: dict[str, list[str]] = {
    "Venezuelan State-Owned Oil Company (PDVSA)": [
        "PDVSA", "Petroleos de Venezuela", "Petróleos de Venezuela SA",
    ],
    "State Development Bank VEB.RF": [
        "Vnesheconombank", "VEB.RF", "VEB Bank",
        "Bank for Development and Foreign Economic Affairs",
    ],
    "Belnauchcompositit": ["Belnauchkomposit", "Belnaukcomposit"],
    "Belorusskaya Kaliynaya Companya": [
        "Belarusian Potash Company", "Belarusian Potash Corporation", "BPC",
    ],
}

# Column-name hints -> canonical field. Matching is case-insensitive substring.
COLUMN_HINTS = {
    "name": ["name", "entity", "company", "supplier", "vendor", "counterparty"],
    "address": ["address", "street", "location"],
    "country": ["country", "nation", "jurisdiction"],
    "type": ["type", "entity_type", "kind"],
    "identifier": ["identifier", "id_number", "reg", "tax", "duns"],
}


@dataclass
class InputEntity:
    row: int
    name: str
    address: str | None = None
    country: str | None = None
    type: str | None = None
    identifier: str | None = None


def _detect_columns(header: list[str]) -> dict[str, int]:
    """Map canonical field -> column index using COLUMN_HINTS."""
    mapping: dict[str, int] = {}
    lowered = [(i, str(h).strip().lower()) for i, h in enumerate(header) if h is not None]
    for field_name, hints in COLUMN_HINTS.items():
        for i, h in lowered:
            if any(hint in h for hint in hints):
                mapping[field_name] = i
                break
    return mapping


def load_entities_from_xlsx(path: Path, sheet: str = "list_1") -> list[InputEntity]:
    """Read the entity list. Auto-detects columns; ``name`` is required."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.active
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    cols = _detect_columns(header)
    if "name" not in cols:
        raise ValueError(
            f"Could not find a name column in {header!r}. "
            "Rename the column to include 'name' or edit COLUMN_HINTS."
        )

    def cell(values: tuple, key: str) -> str | None:
        idx = cols.get(key)
        if idx is None or idx >= len(values):
            return None
        v = values[idx]
        return str(v).strip() if v not in (None, "") else None

    entities: list[InputEntity] = []
    for n, values in enumerate(rows, start=2):  # row 1 is the header
        name = cell(values, "name")
        if not name:
            continue
        entities.append(
            InputEntity(
                row=n,
                name=name,
                address=cell(values, "address"),
                country=cell(values, "country"),
                type=cell(values, "type"),
                identifier=cell(values, "identifier"),
            )
        )
    return entities


# ---------------------------------------------------------------------------
# Steps 2-4: resolve, fetch, extract  (the Sayari API calls)
# ---------------------------------------------------------------------------

@dataclass
class Profile:
    """Flat, traceable view of one entity. Every field cites its source call."""
    input_row: int
    input_name: str
    matched: bool
    entity_id: str | None = None
    match_label: str | None = None
    match_score: float | None = None
    name_mismatch_flag: bool = False  # True if no input-name word appears in matched label
    retry_name_used: str | None = None  # set when a fallback name was needed to resolve
    type: str | None = None
    countries: list[str] = field(default_factory=list)
    sanctioned: bool | None = None
    pep: bool | None = None
    risk_factors: list[str] = field(default_factory=list)
    degree: int | None = None
    relationship_counts: dict[str, int] = field(default_factory=dict)
    source_count: int | None = None
    entity_url: str | None = None
    error: str | None = None


def resolve_entity(client, e: InputEntity) -> tuple[dict | None, dict]:
    """Call the resolution endpoint. Returns (best_match_dict, raw_response).

    Resolution is tuned to return the best match, so we take data[0] but keep
    the score so weak matches can be flagged downstream.
    """
    kwargs: dict[str, Any] = {"name": [e.name]}
    if e.address:
        kwargs["address"] = [e.address]
    if e.country:
        kwargs["country"] = [e.country]
    if e.type:
        kwargs["type"] = [e.type]

    raw = to_dict(client.resolution.resolution(**kwargs))
    candidates = raw.get("data") or []
    best = candidates[0] if candidates else None
    return best, raw


def resolve_with_fallback(
    client, e: InputEntity
) -> tuple[dict | None, dict, str | None]:
    """Try the primary name first; on no-match, try RETRY_NAMES alternates and
    auto-extracted acronyms (e.g. '...  (PDVSA)' -> 'PDVSA').

    Returns (best_match_dict, raw_resolution, retry_name_used).
    retry_name_used is None if the primary name succeeded or a pin was used.

    PINNED_IDS entries short-circuit the API call entirely and return a synthetic
    match dict so the entity is always fetched from the verified parent ID.
    """
    # Check pin override first — no API call needed for the resolution step
    pinned_id = PINNED_IDS.get(e.name)
    if pinned_id:
        synthetic_match = {"entity_id": pinned_id, "label": f"[pinned] {e.name}", "score": None}
        synthetic_raw = {"pinned": True, "entity_id": pinned_id, "input_name": e.name}
        return synthetic_match, synthetic_raw, f"[pinned:{pinned_id}]"

    match, raw = resolve_entity(client, e)
    if match:
        return match, raw, None

    # Build the list of alternates to try
    alternates: list[str] = list(RETRY_NAMES.get(e.name, []))
    # Auto-extract parenthetical acronym, e.g. "Full Name (ABC)" -> try "ABC"
    m = re.search(r'\(([A-Z]{2,})\)', e.name)
    if m:
        acronym = m.group(1)
        if acronym not in alternates:
            alternates.insert(0, acronym)

    for alt in alternates:
        e_alt = InputEntity(row=e.row, name=alt, country=e.country)
        match, raw = resolve_entity(client, e_alt)
        if match:
            return match, raw, alt

    return None, raw, None


def fetch_profile(client, entity_id: str) -> dict:
    """Call get_entity for the full profile. Returns the raw response as a dict."""
    return to_dict(client.entity.get_entity(id=entity_id))


def extract_profile(e: InputEntity, match: dict | None, raw_entity: dict | None,
                    effective_name: str | None = None) -> Profile:
    """Flatten resolution + entity responses into a traceable Profile record.

    Field names are pulled defensively; the authoritative values live in the
    saved raw/<id>.json. Confirm exact paths there after the first live run.
    """
    p = Profile(input_row=e.row, input_name=e.name, matched=match is not None)
    if match is None:
        return p

    p.entity_id = first_present(match, "entity_id", "id")
    # For pinned IDs the match dict is synthetic; real label comes from entity profile
    raw_label = first_present(match, "label", "name") or ""
    p.match_label = raw_label if not raw_label.startswith("[pinned]") else None
    # resolution score may be nested under match_strength/score depending on schema
    p.match_score = (
        first_present(match, "score")
        or deep_get(match, "match_strength", "value")
    )
    # Mismatch flag: no significant input-name word appears in matched label.
    # Checks: (1) raw label, (2) unidecode transliteration of Cyrillic → Latin,
    # (3) translated_label (English rendering stored in entity profile).
    # effective_name overrides e.name when a retry alternate was used.
    if p.match_label:
        stop_words = {
            "the", "of", "and", "for", "a", "an", "in", "co", "ltd", "llc", "inc",
            "jsc", "ojsc", "pjsc", "pao", "oao", "ooo", "ao", "sa",
        }
        check_name = effective_name or e.name
        input_words = {
            w.lower() for w in check_name.split()
            if w.lower() not in stop_words and len(w) > 2
        }
        label_lower = p.match_label.lower()
        label_ascii = unidecode(p.match_label).lower()   # Cyrillic -> Latin
        # translated_label: English rendering stored in get_entity profile
        translated = ""
        if raw_entity:
            _d = raw_entity.get("data") if isinstance(raw_entity.get("data"), dict) else raw_entity
            translated = (_d.get("translated_label") or "").lower()
        p.name_mismatch_flag = bool(input_words) and not any(
            w in label_lower or w in label_ascii or w in translated
            for w in input_words
        )

    if not raw_entity:
        return p

    data = raw_entity.get("data") if isinstance(raw_entity.get("data"), dict) else raw_entity

    # Fill label from entity profile when it was absent (e.g. pinned ID)
    if not p.match_label:
        p.match_label = first_present(data, "label", "translated_label")

    p.type = first_present(data, "type", "entity_type")
    countries = first_present(data, "countries", "country", default=[])
    p.countries = countries if isinstance(countries, list) else [countries]

    risk = data.get("risk") or {}
    # `sanctioned` and `pep` can live at top level or inside `risk`
    p.sanctioned = first_present(data, "sanctioned") if "sanctioned" in data else risk.get("sanctioned")
    p.pep = first_present(data, "pep") if "pep" in data else risk.get("pep")
    # risk factor names = keys present/truthy in the risk object.
    # Each risk value is a dict {"value": ..., "metadata": ..., "level": ...};
    # we check v["value"] (confirmed against live raw/<id>.json).
    if isinstance(risk, dict):
        p.risk_factors = sorted(
            k for k, v in risk.items()
            if (v.get("value") not in (None, False, 0, "", [], {})
                if isinstance(v, dict)
                else v not in (None, False, 0, "", [], {}))
        )

    p.degree = first_present(data, "degree")
    rc = first_present(data, "relationship_count", "relationship_counts", default={})
    p.relationship_counts = {k: v for k, v in rc.items() if isinstance(v, int)} if isinstance(rc, dict) else {}
    # source_count from get_entity is a dict {source_id: {count, label, ...}};
    # we store the number of distinct sources (confirmed against live raw/<id>.json).
    sc = first_present(data, "source_count", "sources_count")
    p.source_count = (
        len(sc) if isinstance(sc, dict)
        else sc if isinstance(sc, int)
        else len(data["sources"]) if isinstance(data.get("sources"), list)
        else None
    )
    # entity_url is a relative path (/v1/entity/{id}); store as-is.
    p.entity_url = first_present(data, "entity_url", "url", "sayari_url")
    return p


# ---------------------------------------------------------------------------
# Step 5: macro-level aggregation (the actual "analytics report")
# ---------------------------------------------------------------------------

def build_summary(profiles: list[Profile]) -> dict:
    matched = [p for p in profiles if p.matched]
    countries: Counter = Counter()
    for p in matched:
        for c in p.countries:
            if c:
                countries[c] += 1
    risk_factor_freq: Counter = Counter()
    for p in matched:
        risk_factor_freq.update(p.risk_factors)

    sanctioned = sum(1 for p in matched if p.sanctioned)
    pep = sum(1 for p in matched if p.pep)

    return {
        "total_input": len(profiles),
        "resolved": len(matched),
        "unresolved": len(profiles) - len(matched),
        "resolution_rate": round(len(matched) / len(profiles), 3) if profiles else 0,
        "sanctioned_count": sanctioned,
        "pep_count": pep,
        "sanctioned_pct": round(sanctioned / len(matched), 3) if matched else 0,
        "pep_pct": round(pep / len(matched), 3) if matched else 0,
        "country_breakdown": dict(countries.most_common()),
        "entity_type_breakdown": dict(Counter(p.type for p in matched if p.type).most_common()),
        "risk_factor_frequency": dict(risk_factor_freq.most_common()),
        # Resolution score is a raw relevance float (not 0-1 normalized).
        # Confirmed live scores are 100-200+ for strong matches; flag < 50 as low-confidence.
        # Also flag name_mismatch_flag=True: no input-name word appears in matched label
        # (e.g. "Sberbank" -> "Сбербанк-Сервис" — subsidiary match, not parent).
        "low_confidence_matches": [
            {"input_name": p.input_name, "matched": p.match_label, "score": p.match_score,
             "reason": ("low_score" if isinstance(p.match_score, (int, float)) and p.match_score < 50
                        else "name_mismatch")}
            for p in matched
            if (isinstance(p.match_score, (int, float)) and p.match_score < 50)
               or p.name_mismatch_flag
        ],
    }


# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------

def write_outputs(out_dir: Path, profiles: list[Profile], summary: dict) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)

    # resolution_log.csv -- audit which input matched which entity, and how well
    with (out_dir / "resolution_log.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([
            "input_row", "input_name", "matched", "entity_id", "matched_label",
            "match_score", "name_mismatch_flag", "retry_name_used",
        ])
        for p in profiles:
            w.writerow([
                p.input_row, p.input_name, p.matched, p.entity_id, p.match_label,
                p.match_score, p.name_mismatch_flag, p.retry_name_used,
            ])

    # entities.csv -- the flat ground-truth table to diff against the agent
    with (out_dir / "entities.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([
            "entity_id", "input_name", "matched_label", "type", "countries",
            "sanctioned", "pep", "risk_factors", "degree", "source_count", "entity_url",
        ])
        for p in profiles:
            if not p.matched:
                continue
            w.writerow([
                p.entity_id, p.input_name, p.match_label, p.type, "; ".join(p.countries),
                p.sanctioned, p.pep, "; ".join(p.risk_factors), p.degree, p.source_count, p.entity_url,
            ])

    # summary.json -- macro insights for the report
    (out_dir / "summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def build_client():
    """Instantiate the official Sayari SDK client (handles OAuth + refresh)."""
    cid = os.environ.get("SAYARI_CLIENT_ID")
    secret = os.environ.get("SAYARI_CLIENT_SECRET")
    if not cid or not secret:
        sys.exit("Set SAYARI_CLIENT_ID and SAYARI_CLIENT_SECRET environment variables.")
    from sayari.client import Sayari
    return Sayari(client_id=cid, client_secret=secret)


def main() -> None:
    ap = argparse.ArgumentParser(description="Deterministic Sayari reference pipeline.")
    ap.add_argument("--input", required=True, type=Path, help="Path to the entity .xlsx")
    ap.add_argument("--out", type=Path, default=Path("./output"), help="Output directory")
    ap.add_argument("--limit", type=int, default=0, help="Process only the first N rows (0 = all)")
    ap.add_argument("--sleep", type=float, default=0.2, help="Seconds between API calls (rate limiting)")
    ap.add_argument("--verbose", action="store_true")
    args = ap.parse_args()

    entities = load_entities_from_xlsx(args.input)
    if args.limit:
        entities = entities[: args.limit]
    print(f"Loaded {len(entities)} entities from {args.input}")

    client = build_client()
    raw_dir = args.out / "raw"
    raw_dir.mkdir(parents=True, exist_ok=True)

    profiles: list[Profile] = []
    cache: dict[str, dict] = {}  # entity_id -> raw profile (avoid duplicate calls)

    for e in entities:
        try:
            match, raw_res, retry_name = resolve_with_fallback(client, e)
            # save the resolution payload keyed by input row
            (raw_dir / f"row{e.row}_resolution.json").write_text(
                json.dumps(raw_res, indent=2), encoding="utf-8"
            )
            raw_entity = None
            if match:
                eid = first_present(match, "entity_id", "id")
                if eid:
                    if eid in cache:
                        raw_entity = cache[eid]
                    else:
                        raw_entity = fetch_profile(client, eid)
                        cache[eid] = raw_entity
                        (raw_dir / f"{eid}.json").write_text(
                            json.dumps(raw_entity, indent=2), encoding="utf-8"
                        )
            p = extract_profile(e, match, raw_entity, effective_name=retry_name)
            p.retry_name_used = retry_name
        except Exception as exc:  # keep going; record the failure
            p = Profile(input_row=e.row, input_name=e.name, matched=False, error=str(exc))
            print(f"  ! row {e.row} ({e.name}): {exc}", file=sys.stderr)

        profiles.append(p)
        if args.verbose:
            retry_note = f" [retry: {p.retry_name_used!r}]" if p.retry_name_used else ""
            print(f"  row {e.row}: {e.name!r} -> {p.entity_id} "
                  f"(score={p.match_score}, sanctioned={p.sanctioned}, pep={p.pep}){retry_note}")
        time.sleep(args.sleep)

    summary = build_summary(profiles)
    write_outputs(args.out, profiles, summary)

    print(f"\nDone. {summary['resolved']}/{summary['total_input']} resolved "
          f"({summary['sanctioned_count']} sanctioned, {summary['pep_count']} PEP).")
    print(f"Outputs written to {args.out.resolve()}")
    if summary["low_confidence_matches"]:
        print(f"  ⚠ {len(summary['low_confidence_matches'])} uncertain matches "
              f"(low score or label mismatch) -- review resolution_log.csv before trusting these.")


if __name__ == "__main__":
    main()
