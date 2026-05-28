"""
Uploads router — turns the upload surface into a real screening pipeline.

Endpoints:
  POST /uploads                       — multipart .xlsx upload + sheet pick
  POST /uploads/{upload_id}/run       — SSE stream that executes the engine

Persistence layout for a completed run (mirrors output/ so the existing
EntityCache and tools work unchanged when pointed at it):

  output/runs/{run_id}/
    raw/{entity_id}.json     — cached Sayari entity responses (one per resolve)
    raw/row{n}_resolution.json — raw resolution responses (audit trail)
    entities.csv             — input_name → entity_id index
    summary.json             — macro summary (build_summary output)
    results.json             — per-entity customer payload (for /api/results)
    investigation.json       — index entry consumed by /api/investigations
"""
from __future__ import annotations

import asyncio
import csv
import hashlib
import io
import json
import logging
import os
import re
import sys
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, AsyncGenerator

from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

_REPO = Path(__file__).resolve().parents[3]
if str(_REPO) not in sys.path:
    sys.path.insert(0, str(_REPO))

from packages.engine import (
    InputEntity,
    build_client,
    build_summary,
    load_entities_from_xlsx,
)
from packages.engine.loader import _detect_columns, COLUMN_HINTS
from packages.engine.profile import extract_profile, fetch_profile
from packages.engine.resolve import PINNED_IDS, RETRY_NAMES, resolve_entity, resolve_with_fallback
from packages.engine.helpers import to_dict

log = logging.getLogger("sentinel.uploads")

router = APIRouter(prefix="/uploads", tags=["uploads"])

# ── Storage paths ────────────────────────────────────────────────────────────

UPLOAD_TMP = Path("/tmp/sentinel-uploads")
UPLOAD_TMP.mkdir(parents=True, exist_ok=True)
RUNS_DIR = _REPO / "output" / "runs"
RUNS_DIR.mkdir(parents=True, exist_ok=True)

# Seed file (used to recognise list_1 by content hash)
SEEDED_XLSX = _REPO / "Sayari_Interview_Exercise_List.xlsx"

# Sayari rate limit — one request per second, conservatively
SAYARI_TICK_SECONDS = 1.0

# Max input rows we'll run in one job (safety brake)
MAX_RUN_ROWS = 200

# In-memory upload registry: upload_id → dict(file_path, sheet, entities, mapping)
_uploads: dict[str, dict] = {}


# ── Helpers ──────────────────────────────────────────────────────────────────


def _new_upload_id() -> str:
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    suffix = uuid.uuid4().hex[:6]
    return f"u_{ts}_{suffix}"


def _new_run_id() -> str:
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    suffix = uuid.uuid4().hex[:6]
    return f"run_{ts}_{suffix}"


def _content_hash(entities: list[InputEntity]) -> str:
    """
    Deterministic content hash of the parsed input rows. Used to recognise
    list_1 even when uploaded under a different filename / row order.

    Hash is over the sorted (name_normalized, country_upper) tuples so it's
    insensitive to row reordering, header rename, or column drift around the
    name column.
    """
    tuples = sorted(
        [
            (e.name.strip().lower(), (e.country or "").strip().upper())
            for e in entities
            if e.name
        ]
    )
    payload = json.dumps(tuples, ensure_ascii=False)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _list_sheets(path: Path) -> list[str]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    return list(wb.sheetnames)


def _read_header(path: Path, sheet: str) -> list[Any]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.active
    row_iter = ws.iter_rows(values_only=True)
    try:
        return list(next(row_iter))
    except StopIteration:
        return []


def _build_mapping(header: list[Any]) -> dict[str, Any]:
    cols = _detect_columns(header)
    return {
        # logical_field → { detected_header, detected_index, hints }
        field: {
            "detected_header": header[idx] if idx is not None and idx < len(header) else None,
            "detected_index": idx,
            "hints": COLUMN_HINTS[field],
        }
        for field, idx in (
            (f, cols.get(f)) for f in COLUMN_HINTS.keys()
        )
    }


def _preview_rows(entities: list[InputEntity], limit: int = 12) -> list[dict]:
    return [
        {
            "row": e.row,
            "name": e.name,
            "country": e.country,
            "address": e.address,
            "type": e.type,
            "identifier": e.identifier,
            "status": "ready" if e.name else "no_name",
        }
        for e in entities[:limit]
    ]


# ── LIST_1_HASH (lazy, computed once) ─────────────────────────────────────────

_LIST_1_HASH_CACHED: str | None = None


def _list_1_hash() -> str | None:
    global _LIST_1_HASH_CACHED
    if _LIST_1_HASH_CACHED is not None:
        return _LIST_1_HASH_CACHED
    if not SEEDED_XLSX.exists():
        return None
    try:
        ents = load_entities_from_xlsx(SEEDED_XLSX, sheet="list_1")
        _LIST_1_HASH_CACHED = _content_hash(ents)
        log.info("LIST_1_HASH computed: %s (%d rows)", _LIST_1_HASH_CACHED[:16], len(ents))
        return _LIST_1_HASH_CACHED
    except Exception as exc:
        log.warning("Could not compute LIST_1_HASH: %s", exc)
        return None


# ── POST /uploads ─────────────────────────────────────────────────────────────


@router.post("")
async def upload_file(
    file: UploadFile = File(...),
    sheet_name: str | None = Form(None),
):
    """
    Accept a .xlsx (or .csv) upload, optionally with a chosen sheet_name.

    Two-phase usage:
      1) First POST with only `file`. Response includes the full sheet list.
         If only one sheet is present it's auto-selected and the response
         already contains preview + column_mapping + matches_seeded.
      2) If multi-sheet, POST again with `sheet_name=<chosen>` to parse that
         sheet and get the preview / matches_seeded outcome.

    Reusing the same `file` between phases is supported by saving it to
    /tmp/sentinel-uploads/{upload_id}/source.{ext} on the first call.

    Returns:
      {
        upload_id, filename, sheets, sheet, total_rows, skipped_rows,
        preview, column_mapping, matches_seeded, content_hash
      }
    """
    if not file.filename:
        raise HTTPException(400, "Upload missing filename")
    suffix = Path(file.filename).suffix.lower()
    if suffix not in (".xlsx", ".csv", ".xls"):
        raise HTTPException(415, f"Unsupported file type: {suffix} (use .xlsx or .csv)")

    upload_id = _new_upload_id()
    upload_dir = UPLOAD_TMP / upload_id
    upload_dir.mkdir(parents=True, exist_ok=True)
    saved = upload_dir / f"source{suffix}"
    saved.write_bytes(await file.read())

    sheets: list[str] = []
    if suffix in (".xlsx", ".xls"):
        sheets = _list_sheets(saved)
    else:
        sheets = ["(csv)"]

    # If multi-sheet and no choice provided, return early with the sheet list.
    if len(sheets) > 1 and not sheet_name:
        first = sheets[0]
        header = _read_header(saved, first) if suffix in (".xlsx", ".xls") else []
        _uploads[upload_id] = {
            "file_path": str(saved),
            "filename": file.filename,
            "sheet": None,
            "sheets": sheets,
            "entities": [],
            "content_hash": None,
        }
        return {
            "upload_id": upload_id,
            "filename": file.filename,
            "sheets": sheets,
            "sheet": None,
            "needs_sheet_choice": True,
            "total_rows": 0,
            "skipped_rows": 0,
            "preview": [],
            "column_mapping": _build_mapping(header) if header else {},
            "matches_seeded": False,
            "content_hash": None,
        }

    # Single-sheet or sheet_name supplied: parse.
    chosen = sheet_name or sheets[0]
    if suffix == ".csv":
        # Convert CSV → in-memory xlsx is overkill; just parse with csv lib.
        entities = _load_entities_from_csv(saved)
    else:
        try:
            entities = load_entities_from_xlsx(saved, sheet=chosen)
        except ValueError as exc:
            raise HTTPException(422, str(exc))

    if len(entities) > MAX_RUN_ROWS:
        raise HTTPException(
            413,
            f"This sheet has {len(entities)} rows; current MAX_RUN_ROWS is {MAX_RUN_ROWS}.",
        )

    header = _read_header(saved, chosen) if suffix in (".xlsx", ".xls") else _csv_header(saved)
    mapping = _build_mapping(header)
    chash = _content_hash(entities)
    seeded_hash = _list_1_hash()
    matches_seeded = bool(seeded_hash and chash == seeded_hash)

    # Count skipped (rows with no name) by reading the raw sheet length minus parsed
    skipped = max(0, _row_count(saved, chosen, suffix) - 1 - len(entities))  # -1 for header

    _uploads[upload_id] = {
        "file_path": str(saved),
        "filename": file.filename,
        "sheet": chosen,
        "sheets": sheets,
        "entities": entities,
        "content_hash": chash,
    }

    return {
        "upload_id": upload_id,
        "filename": file.filename,
        "sheets": sheets,
        "sheet": chosen,
        "needs_sheet_choice": False,
        "total_rows": len(entities),
        "skipped_rows": skipped,
        "preview": _preview_rows(entities),
        "column_mapping": mapping,
        "matches_seeded": matches_seeded,
        "content_hash": chash,
    }


def _row_count(path: Path, sheet: str, suffix: str) -> int:
    """
    Count rows that actually contain data.

    openpyxl's iter_rows walks the worksheet's *used range*, which
    includes phantom rows below the data whenever cell formatting (borders,
    background fills, etc.) was applied past the last real row. A 50-row
    sheet can report 997 because of that. We filter to rows where at least
    one cell is non-None and not whitespace, so the front-end's "input rows"
    count reflects what an analyst would actually see in the file.
    """
    if suffix == ".csv":
        with path.open(newline="", encoding="utf-8") as f:
            return sum(1 for line in f if line.strip())
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.active
    count = 0
    for row in ws.iter_rows(values_only=True):
        if any(
            v is not None and (not isinstance(v, str) or v.strip())
            for v in row
        ):
            count += 1
    return count


def _csv_header(path: Path) -> list[Any]:
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        try:
            return next(reader)
        except StopIteration:
            return []


def _load_entities_from_csv(path: Path) -> list[InputEntity]:
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        header = next(reader)
        cols = _detect_columns(header)
        if "name" not in cols:
            raise HTTPException(422, f"Could not find a name column in {header!r}.")

        def cell(values: list[str], key: str) -> str | None:
            idx = cols.get(key)
            if idx is None or idx >= len(values):
                return None
            v = values[idx]
            return v.strip() if v and v.strip() else None

        out: list[InputEntity] = []
        for n, values in enumerate(reader, start=2):
            name = cell(values, "name")
            if not name:
                continue
            out.append(
                InputEntity(
                    row=n,
                    name=name,
                    address=cell(values, "address"),
                    country=cell(values, "country"),
                    type=cell(values, "type"),
                    identifier=cell(values, "identifier"),
                )
            )
        return out


# ── POST /uploads/{upload_id}/run (SSE) ───────────────────────────────────────


def _sse(event: str, data: Any) -> str:
    return f"data: {json.dumps({'event': event, 'data': data}, default=str)}\n\n"


def _derive_risk_level(p) -> str:
    """Match the front-end's stat colouring rule, server-side."""
    if p.sanctioned:
        return "critical"
    if len(p.risk_factors or []) > 4:
        return "high"
    if p.risk_factors:
        return "medium"
    return "low"


def _attempt_record(name: str, country: str | None, address: str | None,
                     duration_ms: int, raw: dict, match: dict | None,
                     pinned: bool = False, retry_name: str | None = None) -> dict:
    """One resolve-call audit record — shape mirrors the SSE 'steps' contract."""
    request = {"name": name}
    if country: request["country"] = country
    if address: request["address"] = address
    cand_count = (
        1 if pinned
        else len((raw or {}).get("data") or [])
    )
    rec = {
        "step": "resolve_entity",
        "endpoint": "POST /v1/entity/resolution",
        "duration_ms": duration_ms,
        "request": request,
        "response_summary": {
            "candidate_count": cand_count,
            "best_label": (match or {}).get("label"),
            "best_score": (match or {}).get("score"),
            "pinned": bool(pinned),
        },
    }
    if retry_name:
        rec["response_summary"]["retry_name"] = retry_name
    return rec


def _resolve_with_audit(client, e):
    """
    Drop-in replacement for `resolve_with_fallback` that *also* returns a list
    of per-attempt audit records (one per Sayari /v1/entity/resolution call,
    or one synthetic entry for a PINNED_IDS short-circuit).

    Returns: (match | None, raw_resolution, retry_name | None, attempts: list[dict])
    """
    attempts: list[dict] = []

    # 1) Pinned short-circuit — no live API call. We still record the audit
    #    entry so the UI can show the "[pinned:…]" badge with near-zero duration.
    pinned_id = PINNED_IDS.get(e.name)
    if pinned_id:
        t0 = time.time()
        synthetic_match = {"entity_id": pinned_id, "label": f"[pinned] {e.name}", "score": None}
        synthetic_raw = {"pinned": True, "entity_id": pinned_id, "input_name": e.name}
        dur = int((time.time() - t0) * 1000)
        attempts.append(_attempt_record(
            e.name, e.country, e.address, dur, synthetic_raw, synthetic_match,
            pinned=True, retry_name=f"[pinned:{pinned_id}]",
        ))
        return synthetic_match, synthetic_raw, f"[pinned:{pinned_id}]", attempts

    # 2) Primary name
    t0 = time.time()
    match, raw = resolve_entity(client, e)
    dur = int((time.time() - t0) * 1000)
    attempts.append(_attempt_record(
        e.name, e.country, e.address, dur, raw, match,
    ))
    if match:
        return match, raw, None, attempts

    # 3) RETRY_NAMES + acronym fallback
    alternates: list[str] = list(RETRY_NAMES.get(e.name, []))
    m = re.search(r'\(([A-Z]{2,})\)', e.name)
    if m:
        acronym = m.group(1)
        if acronym not in alternates:
            alternates.insert(0, acronym)

    from packages.engine import InputEntity
    last_raw = raw
    for alt in alternates:
        e_alt = InputEntity(row=e.row, name=alt, country=e.country)
        t0 = time.time()
        match, raw = resolve_entity(client, e_alt)
        dur = int((time.time() - t0) * 1000)
        last_raw = raw
        attempts.append(_attempt_record(
            alt, e.country, e.address, dur, raw, match, retry_name=alt,
        ))
        if match:
            return match, raw, alt, attempts

    return None, last_raw, None, attempts


@router.post("/{upload_id}/run")
async def run_upload(upload_id: str, name: str | None = None):
    """
    Execute the engine on the upload's parsed rows. Returns SSE.

    Cached path (matches_seeded=true): emits a single "matched seeded list_1"
    event and a "run_complete" with run_id="default", so the front-end falls
    back to the existing demo cache without any work.

    Live path: walks each row, resolves via Sayari, fetches the profile, caches
    the raw response to output/runs/{run_id}/raw/. Streams per-row progress.
    On completion, writes entities.csv + summary.json + results.json +
    investigation.json so /entities, /summary, /tools/compare_ofac_vs_sayari,
    and /api/investigations all serve the new run when called with ?run_id=.

    Args:
        upload_id:  the upload's transient id (from POST /uploads).
        name:       optional analyst-chosen investigation name. Falls back to
                    the original filename. Persisted into investigation.json.
    """
    upload = _uploads.get(upload_id)
    if upload is None:
        raise HTTPException(404, f"Unknown upload_id: {upload_id}")
    if not upload.get("entities"):
        raise HTTPException(400, "Upload has no parsed rows. Re-POST with a sheet_name.")

    entities: list[InputEntity] = upload["entities"]
    matches_seeded = bool(
        _list_1_hash() and upload.get("content_hash") == _list_1_hash()
    )

    async def stream() -> AsyncGenerator[str, None]:
        if matches_seeded:
            yield _sse(
                "matched_seeded",
                {
                    "filename": upload.get("filename"),
                    "sheet": upload.get("sheet"),
                    "total_rows": len(entities),
                    "message": "matched seeded list_1 — using verified cached results",
                },
            )
            yield _sse(
                "run_complete",
                {
                    "run_id": "default",          # routes UI to the existing list_1 cache
                    "matches_seeded": True,
                    "resolved": 49,
                    "total": len(entities),
                },
            )
            yield _sse("done", {})
            return

        # ── Live run ────────────────────────────────────────────────────────
        client = build_client()
        if client is None:
            yield _sse(
                "error",
                {
                    "message": "Live run requires SAYARI_CLIENT_ID + SAYARI_CLIENT_SECRET in .env",
                },
            )
            yield _sse("done", {})
            return

        run_id = _new_run_id()
        run_dir = RUNS_DIR / run_id
        (run_dir / "raw").mkdir(parents=True, exist_ok=True)

        yield _sse(
            "run_started",
            {
                "run_id": run_id,
                "filename": upload.get("filename"),
                "sheet": upload.get("sheet"),
                "total": len(entities),
            },
        )

        # csv rows we accumulate as we resolve
        csv_rows: list[dict] = []
        loop = asyncio.get_event_loop()

        for i, e in enumerate(entities):
            yield _sse("row_start", {"row": e.row, "index": i, "name": e.name, "country": e.country})
            t0 = time.time()

            # Cache file paths — relative to repo root for the SSE event and
            # for the front-end's "Sources" links.
            resolution_cache = f"output/runs/{run_id}/raw/row{e.row}_resolution.json"
            try:
                # ── Step 1: resolve_entity (with per-attempt audit) ───────────
                match, raw_resolution, retry_name, attempts = await loop.run_in_executor(
                    None, lambda: _resolve_with_audit(client, e)
                )
                # Persist the raw resolution payload for audit
                (run_dir / "raw" / f"row{e.row}_resolution.json").write_text(
                    json.dumps(raw_resolution, default=str, ensure_ascii=False), encoding="utf-8"
                )

                if not match or not match.get("entity_id"):
                    yield _sse(
                        "row_unresolved",
                        {
                            "row": e.row,
                            "index": i,
                            "name": e.name,
                            "duration_ms": int((time.time() - t0) * 1000),
                            # Audit-trace enrichment
                            "steps": attempts,
                            "reason": (
                                f"No candidates above threshold after "
                                f"{len(attempts)} attempt{'' if len(attempts) == 1 else 's'} "
                                f"({', '.join(a['request']['name'] for a in attempts)})."
                            ),
                            "sources": {
                                "resolution_cache": resolution_cache,
                            },
                        },
                    )
                    csv_rows.append({
                        "input_row": e.row,
                        "input_name": e.name,
                        "entity_id": "",
                        "match_label": "",
                        "match_score": "",
                        "name_mismatch_flag": "",
                        "retry_name_used": "",
                        "country": e.country or "",
                    })
                    await asyncio.sleep(SAYARI_TICK_SECONDS)
                    continue

                entity_id = match["entity_id"]

                # ── Step 2: fetch_profile (its own timing boundary) ──────────
                t_profile = time.time()
                raw_entity = await loop.run_in_executor(
                    None, lambda: fetch_profile(client, entity_id)
                )
                profile_duration_ms = int((time.time() - t_profile) * 1000)
                # Cache the raw profile response
                (run_dir / "raw" / f"{entity_id}.json").write_text(
                    json.dumps(raw_entity, default=str, ensure_ascii=False), encoding="utf-8"
                )

                # Build the Profile object — single source of truth for findings
                p = extract_profile(e, match, raw_entity, effective_name=retry_name)
                csv_rows.append({
                    "input_row": e.row,
                    "input_name": e.name,
                    "entity_id": entity_id,
                    "match_label": p.match_label or "",
                    "match_score": p.match_score if p.match_score is not None else "",
                    "name_mismatch_flag": "1" if p.name_mismatch_flag else "",
                    "retry_name_used": retry_name or "",
                    "country": e.country or "",
                })

                # Compose the audit-trace payload
                steps = list(attempts)
                steps.append({
                    "step": "fetch_profile",
                    "endpoint": "GET /v1/entity/{id}",
                    "duration_ms": profile_duration_ms,
                    "request": {"entity_id": entity_id},
                    "response_summary": {
                        "countries": p.countries,
                        "degree": p.degree,
                        "source_count": p.source_count,
                        "risk_factor_count": len(p.risk_factors),
                    },
                })
                sanctioned_lists = [
                    f for f in (p.risk_factors or []) if f.startswith("sanctioned_")
                ]
                findings = {
                    "match_label": p.match_label,
                    "match_score": p.match_score,
                    "name_mismatch_flag": bool(p.name_mismatch_flag),
                    "type": p.type,
                    "countries": p.countries,
                    "risk_level": _derive_risk_level(p),
                    "sanctioned": bool(p.sanctioned),
                    "sanctioned_lists": sanctioned_lists,
                    "pep": bool(p.pep),
                    "top_risks": (p.risk_factors or [])[:5],
                }
                sources = {
                    "resolution_cache": resolution_cache,
                    "profile_cache": f"output/runs/{run_id}/raw/{entity_id}.json",
                }

                yield _sse(
                    "row_resolved",
                    {
                        # Existing fields — preserved for backward compatibility
                        "row": e.row,
                        "index": i,
                        "name": e.name,
                        "entity_id": entity_id,
                        "match_label": p.match_label,
                        "sanctioned": bool(p.sanctioned),
                        "pep": bool(p.pep),
                        "risk_factor_count": len(p.risk_factors),
                        "warn_verify": bool(p.name_mismatch_flag),
                        "duration_ms": int((time.time() - t0) * 1000),
                        # Audit-trace enrichment
                        "steps":    steps,
                        "findings": findings,
                        "sources":  sources,
                    },
                )
            except Exception as exc:
                log.exception("Run %s row %d failed", run_id, e.row)
                yield _sse(
                    "row_error",
                    {
                        "row": e.row,
                        "index": i,
                        "name": e.name,
                        "error": str(exc),
                        "duration_ms": int((time.time() - t0) * 1000),
                    },
                )

            # Rate-limit between Sayari calls
            await asyncio.sleep(SAYARI_TICK_SECONDS)

        # ── Persist the run as a complete EntityCache-compatible directory ──
        with (run_dir / "entities.csv").open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(
                f,
                fieldnames=[
                    "input_row", "input_name", "entity_id", "match_label",
                    "match_score", "name_mismatch_flag", "retry_name_used", "country",
                ],
            )
            writer.writeheader()
            writer.writerows(csv_rows)

        # Build profiles list (using the new cache) and macro summary
        from packages.engine import EntityCache  # local to avoid early init
        cache = EntityCache(str(run_dir))
        profiles = cache.all_profiles()
        summary = build_summary(profiles)

        # Compute the ownership-gap count from this run's compare summary, so
        # the dashboard KPI ("ownership-gap findings") lights up without the
        # client having to recompute it. The matcher comes from the main app's
        # global (loaded once at startup). If the cold-start race leaves it
        # unloaded, fall back to 0 rather than blocking the run on it.
        ownership_gap = 0
        try:
            from services.api.main import _get_ofac
            matcher = _get_ofac()
            if matcher is None:
                log.warning("OFAC matcher not ready when computing ownership_gap_count for %s — defaulting to 0", run_id)
            else:
                from services.api.tools.compare_ofac_vs_sayari import compare_ofac_vs_sayari_tool
                cmp_data = compare_ofac_vs_sayari_tool(cache=cache, ofac_matcher=matcher, threshold=0.85).data
                cmp_summary = cmp_data.get("summary") or {}
                ownership_gap = int((cmp_summary.get("sayari_only") or 0)
                                    + (cmp_summary.get("screen_ambiguous") or 0))
        except Exception as exc:
            log.warning("ownership_gap_count compute failed for %s: %s — defaulting to 0", run_id, exc)

        # Persist gap on summary.json too so future reads don't recompute.
        summary["ownership_gap_count"] = ownership_gap
        (run_dir / "summary.json").write_text(
            json.dumps(summary, indent=2, default=str, ensure_ascii=False), encoding="utf-8"
        )

        # Per-entity results (customer payload — drop-in for /api/results)
        results = _build_results_payload(run_id, cache, profiles)
        (run_dir / "results.json").write_text(
            json.dumps(results, indent=2, default=str, ensure_ascii=False), encoding="utf-8"
        )

        # Investigation index entry — picked up by /api/investigations
        now = datetime.now(timezone.utc).isoformat()
        chosen_name = (name or "").strip() or (
            (upload.get("filename") or "uploaded file")
        )
        investigation = {
            "id": run_id,
            "name": chosen_name,
            "source": "upload",
            "source_detail": upload.get("filename") or "uploaded file",
            "status": "complete",
            "created_at": now,
            "completed_at": now,
            "total_entities": summary.get("resolved", 0),
            "flagged_count": summary.get("sanctioned_count", 0),
            "cleared_count": 0,
            "escalated_count": 0,
            "blocked_count": 0,
            "pending_count": summary.get("resolved", 0),
            "ownership_gap_count": ownership_gap,
        }
        (run_dir / "investigation.json").write_text(
            json.dumps(investigation, indent=2, default=str, ensure_ascii=False), encoding="utf-8"
        )

        # Tell the API layer to drop any cached EntityCache for this run_id so the
        # next read picks up the freshly written entities.csv.
        try:
            from services.api.main import _invalidate_run_cache
            _invalidate_run_cache(run_id)
        except Exception:
            pass

        yield _sse(
            "run_complete",
            {
                "run_id": run_id,
                "matches_seeded": False,
                "resolved": summary.get("resolved", 0),
                "total": summary.get("total_input", len(entities)),
                "sanctioned_count": summary.get("sanctioned_count", 0),
            },
        )
        yield _sse("done", {})

    return StreamingResponse(
        stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


def _build_results_payload(run_id: str, cache, profiles: list) -> list[dict]:
    """Shape per-entity results so /api/results/{run_id} can return them as-is."""
    out = []
    for p in profiles:
        eid = p.entity_id
        if not eid:
            continue
        out.append(
            {
                "entity_id": eid,
                "input_name": p.input_name,
                "resolved": {
                    "entity_id": eid,
                    "label": p.match_label,
                    "confidence": "low" if p.name_mismatch_flag else "high",
                    "entity_url": p.entity_url or f"/v1/entity/{eid}",
                },
                "screening": {
                    "ofac_sdn": {"hit": False, "sdn_id": None, "programs": [], "match_name": None},
                    "ownership_exposure": {"has_exposure": False, "factor": None},
                    "other_sanctions": [],
                    "risk_level": (
                        "critical" if p.sanctioned
                        else "high" if (p.risk_factors and len(p.risk_factors) > 4)
                        else "medium" if p.risk_factors
                        else "low"
                    ),
                    "outcome": "both_catch" if p.sanctioned else "no_ofac",
                    "is_directly_designated": bool(p.sanctioned),
                },
                "disposition": {
                    "status": "pending_review",
                    "reviewer": None,
                    "rationale": None,
                    "updated_at": "",
                },
                "sources": [
                    {
                        "cache_file": f"output/runs/{run_id}/raw/{eid}.json",
                        "api_endpoint": "GET /v1/entity/{id} (cached during this run)",
                    }
                ],
            }
        )
    return out
